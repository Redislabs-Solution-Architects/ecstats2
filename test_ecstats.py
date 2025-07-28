import pytest
import datetime
import configparser
import boto3
from unittest.mock import Mock, patch
import os
import sys
import tempfile
import openpyxl

# Import the module under test
import ecstats


def create_paginator_side_effect(clusters=None, reserved_instances=None):
    """Shared helper to create paginator side effect with custom data."""
    if clusters is None:
        clusters = []
    if reserved_instances is None:
        reserved_instances = []

    def get_paginator_side_effect(paginator_name):
        mock_paginator = Mock()
        if paginator_name == "describe_cache_clusters":
            mock_paginator.paginate.return_value = [{"CacheClusters": clusters}]
        elif paginator_name == "describe_reserved_cache_nodes":
            mock_paginator.paginate.return_value = [
                {"ReservedCacheNodes": reserved_instances}
            ]
        return mock_paginator

    return get_paginator_side_effect


class TestMetricDefinitions:
    """Test metric definition functions."""

    def test_get_max_metrics_hourly(self):
        """Test hourly metrics definition."""
        metrics = ecstats.get_max_metrics_hourly()

        assert isinstance(metrics, list)
        assert len(metrics) > 0

        # Check structure of metrics
        for metric in metrics:
            assert len(metric) == 3
            metric_name, aggregation, period = metric
            assert isinstance(metric_name, str)
            assert aggregation == "Maximum"
            assert period == ecstats.SECONDS_IN_HOUR

    def test_get_max_metrics_weekly(self):
        """Test weekly metrics definition."""
        metrics = ecstats.get_max_metrics_weekly()

        assert isinstance(metrics, list)
        assert len(metrics) > 0

        # Check structure of metrics
        for metric in metrics:
            assert len(metric) == 3
            metric_name, aggregation, period = metric
            assert isinstance(metric_name, str)
            assert aggregation == "Maximum"
            assert (
                period == ecstats.SECONDS_IN_DAY * ecstats.METRIC_COLLECTION_PERIOD_DAYS
            )


class TestUtilityFunctions:
    """Test utility functions."""

    def test_calc_expiry_time(self):
        """Test expiry time calculation."""
        # Test future date
        future_date = datetime.datetime.utcnow() + datetime.timedelta(days=30)
        future_date = future_date.replace(tzinfo=datetime.timezone.utc)

        days_until_expiry = ecstats.calc_expiry_time(future_date)
        assert 29 <= days_until_expiry <= 30  # Allow for small timing differences

        # Test past date
        past_date = datetime.datetime.utcnow() - datetime.timedelta(days=10)
        past_date = past_date.replace(tzinfo=datetime.timezone.utc)

        days_until_expiry = ecstats.calc_expiry_time(past_date)
        assert days_until_expiry < 0


class TestClusterInfo:
    """Test cluster information retrieval."""

    @patch("boto3.Session")
    def test_get_clusters_info_basic_structure(self, mock_session):
        """Test basic structure of get_clusters_info return value."""
        # Mock the session and clients
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        # Use helper to create paginator side effect
        clusters = [
            {
                "CacheClusterId": "test-cluster-001",
                "CacheClusterStatus": "available",
                "Engine": "redis",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            }
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(clusters)
        )

        # Mock describe_snapshots
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        assert "elc_running_instances" in result
        assert "elc_reserved_instances" in result
        assert "snapshots" in result
        assert isinstance(result["elc_running_instances"], dict)
        assert isinstance(result["elc_reserved_instances"], dict)
        assert isinstance(result["snapshots"], dict)

    @patch("boto3.Session")
    def test_get_clusters_info_redis_engine_only(self, mock_session):
        """Test that Redis engine clusters are correctly included."""
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        clusters = [
            {
                "CacheClusterId": "redis-cluster-001",
                "CacheClusterStatus": "available",
                "Engine": "redis",
                "CacheNodeType": "cache.r6g.large",
                "CacheNodes": [
                    {"CacheNodeId": "0001"},
                    {"CacheNodeId": "0002"},
                ],
            },
            {
                "CacheClusterId": "redis-cluster-002",
                "CacheClusterStatus": "available",
                "Engine": "redis",
                "CacheNodeType": "cache.t3.medium",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(clusters)
        )
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        assert len(result["elc_running_instances"]) == 2
        assert "redis-cluster-001" in result["elc_running_instances"]
        assert "redis-cluster-002" in result["elc_running_instances"]

        # Verify Redis engine is preserved
        for cluster_id, cluster_info in result["elc_running_instances"].items():
            assert cluster_info["Engine"] == "redis"

    @patch("boto3.Session")
    def test_get_clusters_info_valkey_engine_only(self, mock_session):
        """Test that Valkey engine clusters are correctly included."""
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        clusters = [
            {
                "CacheClusterId": "valkey-cluster-001",
                "CacheClusterStatus": "available",
                "Engine": "valkey",
                "CacheNodeType": "cache.r7g.xlarge",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "valkey-cluster-002",
                "CacheClusterStatus": "available",
                "Engine": "valkey",
                "CacheNodeType": "cache.m6g.large",
                "CacheNodes": [
                    {"CacheNodeId": "0001"},
                    {"CacheNodeId": "0002"},
                    {"CacheNodeId": "0003"},
                ],
            },
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(clusters)
        )
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        assert len(result["elc_running_instances"]) == 2
        assert "valkey-cluster-001" in result["elc_running_instances"]
        assert "valkey-cluster-002" in result["elc_running_instances"]

        # Verify Valkey engine is preserved
        for cluster_info in result["elc_running_instances"].values():
            assert cluster_info["Engine"] == "valkey"

    @patch("boto3.Session")
    def test_get_clusters_info_filters_redis_valkey_only(self, mock_session):
        """Test that only Redis and Valkey engines are included, other engines filtered out."""
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        clusters = [
            {
                "CacheClusterId": "redis-cluster",
                "CacheClusterStatus": "available",
                "Engine": "redis",
                "CacheNodeType": "cache.r6g.large",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "valkey-cluster",
                "CacheClusterStatus": "available",
                "Engine": "valkey",
                "CacheNodeType": "cache.m6g.medium",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "memcached-cluster",
                "CacheClusterStatus": "available",
                "Engine": "memcached",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(clusters)
        )
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        # Verify only Redis and Valkey clusters are included
        assert len(result["elc_running_instances"]) == 2
        assert "redis-cluster" in result["elc_running_instances"]
        assert "valkey-cluster" in result["elc_running_instances"]
        assert "memcached-cluster" not in result["elc_running_instances"]

        # Verify engines are correctly preserved
        assert result["elc_running_instances"]["redis-cluster"]["Engine"] == "redis"
        assert result["elc_running_instances"]["valkey-cluster"]["Engine"] == "valkey"

    @patch("boto3.Session")
    def test_get_clusters_info_status_filtering(self, mock_session):
        """Test that only 'available' status clusters are included."""
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        clusters = [
            {
                "CacheClusterId": "available-redis",
                "CacheClusterStatus": "available",
                "Engine": "redis",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "creating-redis",
                "CacheClusterStatus": "creating",
                "Engine": "redis",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "deleting-valkey",
                "CacheClusterStatus": "deleting",
                "Engine": "valkey",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
            {
                "CacheClusterId": "available-valkey",
                "CacheClusterStatus": "available",
                "Engine": "valkey",
                "CacheNodeType": "cache.t3.micro",
                "CacheNodes": [{"CacheNodeId": "0001"}],
            },
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(clusters)
        )
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        # Only available clusters should be included
        assert len(result["elc_running_instances"]) == 2
        assert "available-redis" in result["elc_running_instances"]
        assert "available-valkey" in result["elc_running_instances"]
        assert "creating-redis" not in result["elc_running_instances"]
        assert "deleting-valkey" not in result["elc_running_instances"]

    @patch("boto3.Session")
    def test_get_clusters_info_with_reserved_instances(self, mock_session):
        """Test processing of reserved instances for Redis and Valkey."""
        mock_session_instance = Mock()
        mock_session.return_value = mock_session_instance

        mock_elasticache_client = Mock()
        mock_session_instance.client.return_value = mock_elasticache_client

        reserved_instances = [
            {
                "CacheNodeType": "cache.r6g.large",
                "State": "active",
                "ProductDescription": "redis",
                "CacheNodeCount": 3,
                "StartTime": datetime.datetime.now() - datetime.timedelta(days=30),
                "Duration": 31536000,  # 1 year in seconds
            },
            {
                "CacheNodeType": "cache.m6g.xlarge",
                "State": "active",
                "ProductDescription": "valkey",
                "CacheNodeCount": 2,
                "StartTime": datetime.datetime.now() - datetime.timedelta(days=60),
                "Duration": 94608000,  # 3 years in seconds
            },
            {
                "CacheNodeType": "cache.t3.micro",
                "State": "retired",
                "ProductDescription": "redis",
                "CacheNodeCount": 1,
                "StartTime": datetime.datetime.now() - datetime.timedelta(days=400),
                "Duration": 31536000,
            },
            {
                "CacheNodeType": "cache.r5.large",
                "State": "active",
                "ProductDescription": "memcached",
                "CacheNodeCount": 2,
                "StartTime": datetime.datetime.now() - datetime.timedelta(days=30),
                "Duration": 31536000,
            },
        ]
        mock_elasticache_client.get_paginator.side_effect = (
            create_paginator_side_effect(
                clusters=[], reserved_instances=reserved_instances
            )
        )
        mock_elasticache_client.describe_snapshots.return_value = {"Snapshots": []}

        result = ecstats.get_clusters_info(mock_session_instance)

        # Should only include active Redis and Valkey reserved instances
        assert len(result["elc_reserved_instances"]) == 2
        assert "cache.r6g.large" in result["elc_reserved_instances"]
        assert "cache.m6g.xlarge" in result["elc_reserved_instances"]
        assert "cache.t3.micro" not in result["elc_reserved_instances"]  # retired
        assert "cache.r5.large" not in result["elc_reserved_instances"]  # memcached

        # Verify reserved instance details
        redis_ri = result["elc_reserved_instances"]["cache.r6g.large"]
        valkey_ri = result["elc_reserved_instances"]["cache.m6g.xlarge"]

        assert redis_ri["count"] == 3
        assert valkey_ri["count"] == 2
        assert isinstance(redis_ri["expiry_time"], int)
        assert isinstance(valkey_ri["expiry_time"], int)


class TestMetricRetrieval:
    """Test metric retrieval functions."""

    @patch("datetime.date")
    def test_get_metric(self, mock_date):
        """Test get_metric function."""
        # Mock date.today()
        mock_today = datetime.date(2023, 1, 8)
        mock_date.today.return_value = mock_today

        mock_cloudwatch = Mock()
        mock_cloudwatch.get_metric_statistics.return_value = {
            "Datapoints": [{"Maximum": 100.0}, {"Maximum": 150.0}, {"Maximum": 120.0}]
        }

        result = ecstats.get_metric(
            mock_cloudwatch, "test-cluster", "0001", "CurrItems", "Maximum", 3600
        )

        assert result == [100.0, 150.0, 120.0]

        # Verify the CloudWatch call
        mock_cloudwatch.get_metric_statistics.assert_called_once()
        call_args = mock_cloudwatch.get_metric_statistics.call_args

        assert call_args[1]["Namespace"] == "AWS/ElastiCache"
        assert call_args[1]["MetricName"] == "CurrItems"
        assert call_args[1]["Statistics"] == ["Maximum"]

    def test_get_metric_curr(self):
        """Test get_metric_curr function."""
        mock_cloudwatch = Mock()
        mock_cloudwatch.get_metric_data.return_value = {
            "MetricDataResults": [{"Values": [1.0]}]
        }

        result = ecstats.get_metric_curr(
            mock_cloudwatch, "test-cluster", "0001", "IsMaster"
        )

        assert result == 1.0

        # Test empty response
        mock_cloudwatch.get_metric_data.return_value = {
            "MetricDataResults": [{"Values": []}]
        }

        result = ecstats.get_metric_curr(
            mock_cloudwatch, "test-cluster", "0001", "IsMaster"
        )

        assert result == -1


class TestWorkbookOperations:
    """Test Excel workbook operations."""

    def test_create_workbook(self):
        """Test workbook creation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            wb = ecstats.create_workbook(temp_dir, "test-section", "us-west-1")

            assert isinstance(wb, openpyxl.Workbook)
            assert len(wb.sheetnames) == 2
            assert ecstats.RUNNING_INSTANCES_WORKSHEET_NAME in wb.sheetnames
            assert ecstats.RESERVED_INSTANCES_WORKSHEET_NAME in wb.sheetnames

            # Check running instances worksheet headers
            ws = wb[ecstats.RUNNING_INSTANCES_WORKSHEET_NAME]
            headers = [cell.value for cell in ws[1]]

            expected_base_headers = [
                "Source",
                "ClusterId",
                "NodeId",
                "NodeRole",
                "NodeType",
                "Region",
                "SnapshotRetentionLimit",
            ]

            for header in expected_base_headers:
                assert header in headers

            # Should have metrics from both weekly and hourly
            assert "Engine" in headers
            assert "QPF" in headers


class TestIntegration:
    """Integration tests."""

    def test_end_to_end_workflow_mock(self):
        """Test end-to-end workflow with comprehensive mocking."""
        with tempfile.TemporaryDirectory() as temp_dir:
            config_file = os.path.join(temp_dir, "test_config.ini")

            # Create test config
            config = configparser.ConfigParser()
            config.add_section("production")
            config.set("production", "aws_access_key_id", "test-key")
            config.set("production", "aws_secret_access_key", "test-secret")
            config.set("production", "region_name", "us-west-1")

            with open(config_file, "w") as f:
                config.write(f)

            # Mock all AWS interactions
            with patch("boto3.Session") as mock_session, patch(
                "sys.argv", ["ecstats.py", "-c", config_file, "-d", temp_dir]
            ):

                # Setup mock session and clients
                mock_session_instance = Mock()
                mock_session.return_value = mock_session_instance

                mock_elasticache_client = Mock()
                mock_cloudwatch_client = Mock()

                def client_side_effect(service_name):
                    if service_name == "elasticache":
                        return mock_elasticache_client
                    elif service_name == "cloudwatch":
                        return mock_cloudwatch_client
                    return Mock()

                mock_session_instance.client.side_effect = client_side_effect

                # Mock ElastiCache responses using helper
                clusters = [
                    {
                        "CacheClusterId": "test-cluster-001",
                        "CacheClusterStatus": "available",
                        "Engine": "redis",
                        "CacheNodeType": "cache.t3.micro",
                        "PreferredAvailabilityZone": "us-west-1a",
                        "CacheNodes": [{"CacheNodeId": "0001"}],
                    }
                ]
                mock_elasticache_client.get_paginator.side_effect = (
                    create_paginator_side_effect(clusters)
                )

                mock_elasticache_client.describe_snapshots.return_value = {
                    "Snapshots": []
                }

                # Mock CloudWatch responses
                mock_cloudwatch_client.get_metric_statistics.return_value = {
                    "Datapoints": [{"Maximum": 100.0}]
                }
                mock_cloudwatch_client.get_metric_data.return_value = {
                    "MetricDataResults": [{"Values": [1.0]}]
                }

                # Run main function
                ecstats.main()

                # Verify output file was created
                expected_output = os.path.join(temp_dir, "production-us-west-1.xlsx")
                assert os.path.exists(expected_output)

                # Verify the Excel file structure
                wb = openpyxl.load_workbook(expected_output)
                assert ecstats.RUNNING_INSTANCES_WORKSHEET_NAME in wb.sheetnames
                assert ecstats.RESERVED_INSTANCES_WORKSHEET_NAME in wb.sheetnames


if __name__ == "__main__":
    pytest.main([__file__])
